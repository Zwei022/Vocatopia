"""
匯出 Vocatopia 意見回饋成 Excel 檔案。

用法：
    python server/scripts/export_feedback.py

需要環境變數 SUPABASE_URL、SUPABASE_SECRET_KEY（跟後端 server/db/supabase.js
用的是同一組，可以從 `railway variables` 拿，或直接貼在下面兩個常數裡本機執行）。
每次執行都會把「目前資料庫裡所有回饋」重新匯出成一份新檔案，存到桌面，
檔名帶時間戳記，不會覆蓋舊檔案。
"""
import os
import sys
from datetime import datetime

import requests
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

SUPABASE_URL = os.environ.get("SUPABASE_URL", "")
SUPABASE_SECRET_KEY = os.environ.get("SUPABASE_SECRET_KEY", "")
OUT_DIR = os.path.join(os.path.expanduser("~"), "Desktop")


def fetch_feedback():
    if not SUPABASE_URL or not SUPABASE_SECRET_KEY:
        print("缺少 SUPABASE_URL / SUPABASE_SECRET_KEY 環境變數。")
        print("可以先執行：railway run --service Vocatopia python server/scripts/export_feedback.py")
        print("或手動 export SUPABASE_URL=... 、 SUPABASE_SECRET_KEY=... 後再跑一次。")
        sys.exit(1)

    resp = requests.get(
        f"{SUPABASE_URL}/rest/v1/feedback",
        headers={
            "apikey": SUPABASE_SECRET_KEY,
            "Authorization": f"Bearer {SUPABASE_SECRET_KEY}",
        },
        params={"select": "*", "order": "created_at.desc"},
        timeout=30,
    )
    resp.raise_for_status()
    return resp.json()


def export_to_excel(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "意見回饋"

    headers = ["送出時間", "使用者暱稱", "使用者 ID", "內容"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in rows:
        ws.append([
            row.get("created_at", ""),
            row.get("username") or "(未知)",
            row.get("user_id", ""),
            row.get("message", ""),
        ])

    widths = [22, 16, 38, 80]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = os.path.join(OUT_DIR, f"vocatopia_feedback_{timestamp}.xlsx")
    wb.save(out_path)
    return out_path


if __name__ == "__main__":
    rows = fetch_feedback()
    print(f"抓到 {len(rows)} 筆意見回饋")
    path = export_to_excel(rows)
    print(f"已匯出：{path}")
